"""
report.py — Generate CSV summary tables and a combined Excel workbook
=====================================================================
Outputs:
  - summary_stats.csv        — top-level numbers
  - authors_top.csv          — top authors
  - journals_top.csv         — top journals
  - countries_top.csv        — top countries
  - keywords_top.csv         — top keywords
  - mesh_top.csv             — top MeSH terms
  - institutions_top.csv     — top institutions
  - temporal.csv             — year-by-year
  - keratoconus_bibliometrics.xlsx   — all of the above as separate sheets
"""

import csv
import json
import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).parent))
import config

OUT = pathlib.Path(config.OUTPUT_DIR)
OUT.mkdir(parents=True, exist_ok=True)


def _write_csv(filename: str, fieldnames: list, rows: list[dict]):
    path = OUT / filename
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote: {path.name}")


def generate_reports(results: dict):
    n = results["n_records"]

    # ── Summary stats ──────────────────────────────────────────────────────
    temporal = results["temporal"]
    first_year = temporal["years"][0] if temporal["years"] else config.START_YEAR
    last_year  = temporal["years"][-1] if temporal["years"] else config.END_YEAR
    peak_year  = temporal["years"][temporal["counts"].index(max(temporal["counts"]))] \
                 if temporal["counts"] else "N/A"
    total_cites = sum(temporal["citations"]) if temporal["citations"] else 0

    summary_rows = [
        {"metric": "Total publications",          "value": n},
        {"metric": "Year range",                  "value": f"{first_year}–{last_year}"},
        {"metric": "Peak publication year",        "value": peak_year},
        {"metric": "Total citations (CrossRef)",   "value": total_cites},
        {"metric": "Unique journals",              "value": len(results["journals"])},
        {"metric": "Unique countries",             "value": len([c for c in results["countries"] if c["country"] != "Unknown"])},
        {"metric": "Unique institutions (top)",    "value": len(results["institutions"])},
        {"metric": "Unique authors (≥min pubs)",   "value": len(results["authors"])},
        {"metric": "Unique keywords",              "value": len(results["keywords"]["freq"])},
        {"metric": "Unique MeSH terms",            "value": len(results["mesh"]["freq"])},
    ]
    _write_csv("summary_stats.csv", ["metric", "value"], summary_rows)

    # ── Temporal ───────────────────────────────────────────────────────────
    years   = temporal["years"]
    counts  = temporal["counts"]
    cumul   = temporal["cumulative"]
    mavg    = temporal["moving_avg"]
    cites_y = temporal["citations"]

    temporal_rows = []
    for i, y in enumerate(years):
        temporal_rows.append({
            "year":                y,
            "publications":        counts[i],
            "cumulative":          cumul[i],
            "moving_avg_3yr":      round(mavg[i], 1),
            "total_citations":     cites_y[i] if cites_y else 0,
        })
    _write_csv("temporal.csv",
               ["year", "publications", "cumulative", "moving_avg_3yr", "total_citations"],
               temporal_rows)

    # ── Authors ────────────────────────────────────────────────────────────
    author_rows = []
    for rank, a in enumerate(results["authors"][:100], 1):
        author_rows.append({
            "rank":               rank,
            "author":             a["author_id"],
            "publications":       a["pub_count"],
            "first_author":       a["first_author_count"],
            "last_author":        a.get("last_author_count", 0),
            "total_citations":    a["citation_total"],
            "h_index_estimate":   a["h_index_est"],
            "year_first":         a.get("year_first", ""),
            "year_last":          a.get("year_last", ""),
            "years_active":       a.get("years_active", ""),
            "journal_count":      a.get("journal_count", ""),
            "sample_affiliation": "; ".join(a.get("affils_sample", [])),
        })
    _write_csv("authors_top.csv",
               ["rank", "author", "publications", "first_author", "last_author",
                "total_citations", "h_index_estimate", "year_first", "year_last",
                "years_active", "journal_count", "sample_affiliation"],
               author_rows)

    # ── Journals ───────────────────────────────────────────────────────────
    journal_rows = []
    for rank, j in enumerate(results["journals"][:config.TOP_N_JOURNALS], 1):
        journal_rows.append({
            "rank":           rank,
            "journal":        j["journal"],
            "abbreviation":   j["abbr"],
            "publications":   j["count"],
            "percentage":     j["percentage"],
            "total_citations":j["citations"],
        })
    _write_csv("journals_top.csv",
               ["rank", "journal", "abbreviation", "publications",
                "percentage", "total_citations"],
               journal_rows)

    # ── Countries ─────────────────────────────────────────────────────────
    country_rows = []
    valid = [c for c in results["countries"] if c["country"] != "Unknown"]
    for rank, c in enumerate(valid[:config.TOP_N_COUNTRIES], 1):
        country_rows.append({
            "rank":              rank,
            "country":           c["country"],
            "publications":      c["count"],
            "percentage":        c["percentage"],
            "total_citations":   c["citations"],
            "pubs_per_million":  c.get("pubs_per_million", ""),
            "cites_per_million": c.get("cites_per_million", ""),
        })
    _write_csv("countries_top.csv",
               ["rank", "country", "publications", "percentage", "total_citations",
                "pubs_per_million", "cites_per_million"],
               country_rows)

    # ── Keywords ───────────────────────────────────────────────────────────
    kw_freq = results["keywords"]["freq"]
    kw_rows = [{"rank": i+1, "keyword": k, "frequency": v}
               for i, (k, v) in enumerate(
                   sorted(kw_freq.items(), key=lambda x: x[1], reverse=True)[:100]
               )]
    _write_csv("keywords_top.csv", ["rank", "keyword", "frequency"], kw_rows)

    mesh_freq = results["mesh"]["freq"]
    mesh_rows = [{"rank": i+1, "mesh_term": k, "frequency": v}
                 for i, (k, v) in enumerate(
                     sorted(mesh_freq.items(), key=lambda x: x[1], reverse=True)[:100]
                 )]
    _write_csv("mesh_top.csv", ["rank", "mesh_term", "frequency"], mesh_rows)

    # ── Institutions ───────────────────────────────────────────────────────
    inst_rows = [{"rank": i+1, "institution": r["institution"], "publications": r["count"]}
                 for i, r in enumerate(results["institutions"][:50])]
    _write_csv("institutions_top.csv", ["rank", "institution", "publications"], inst_rows)

    # ── Languages ──────────────────────────────────────────────────────────
    _LANG_NAMES = {
        "eng": "English",    "ger": "German",     "rus": "Russian",
        "chi": "Chinese",    "fre": "French",     "rum": "Romanian",
        "heb": "Hebrew",     "cze": "Czech",      "por": "Portuguese",
        "dut": "Dutch",      "hun": "Hungarian",  "pol": "Polish",
        "ita": "Italian",    "slo": "Slovak",     "spa": "Spanish",
        "tur": "Turkish",    "kor": "Korean",     "jpn": "Japanese",
        "ara": "Arabic",     "per": "Persian",
    }
    lang_data = results.get("languages", {})
    lang_rows = [
        {"rank": r, "code": code, "language": _LANG_NAMES.get(code, code), "n": n}
        for r, (code, n) in enumerate(
            sorted(lang_data.items(), key=lambda x: x[1], reverse=True), 1
        )
    ]
    _write_csv("languages.csv", ["rank", "code", "language", "n"], lang_rows)

    # ── Publication types ─────────────────────────────────────────────────
    pub_type_data = results.get("pub_types", {})
    pub_type_rows = [
        {"publication_type": pt, "n": n}
        for pt, n in sorted(pub_type_data.items(), key=lambda x: x[1], reverse=True)
    ]
    _write_csv("pub_types.csv", ["publication_type", "n"], pub_type_rows)


    # ── Excel workbook ─────────────────────────────────────────────────────
    _write_excel(results, summary_rows, temporal_rows, author_rows,
                 journal_rows, country_rows, kw_rows, mesh_rows, inst_rows,
                 lang_rows, pub_type_rows)


def _write_excel(results, summary_rows, temporal_rows, author_rows,
                 journal_rows, country_rows, kw_rows, mesh_rows, inst_rows,
                 lang_rows=None, pub_type_rows=None):
    """Write multi-sheet Excel workbook using openpyxl if available, else skip."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [skip] openpyxl not available — Excel file not created")
        return

    wb = openpyxl.Workbook()

    HEADER_FILL  = PatternFill("solid", fgColor="2C6FAC")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
    TITLE_FONT   = Font(bold=True, size=12)

    def _add_sheet(wb, title, fieldnames, rows, freeze="A2"):
        ws = wb.create_sheet(title=title[:31])
        # Header
        for col, field in enumerate(fieldnames, 1):
            c = ws.cell(row=1, column=col, value=field.replace("_", " ").title())
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        # Rows
        for row_i, row in enumerate(rows, 2):
            fill = ALT_FILL if row_i % 2 == 0 else None
            for col, field in enumerate(fieldnames, 1):
                c = ws.cell(row=row_i, column=col, value=row.get(field, ""))
                if fill:
                    c.fill = fill
                c.alignment = Alignment(vertical="center")
        # Auto-width
        for col in range(1, len(fieldnames) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)
        if freeze:
            ws.freeze_panes = freeze
        return ws

    # Remove default sheet
    del wb["Sheet"]

    _add_sheet(wb, "Summary",      ["metric", "value"],                        summary_rows)
    _add_sheet(wb, "Temporal",     ["year", "publications", "cumulative",
                                     "moving_avg_3yr", "total_citations"],     temporal_rows)
    _add_sheet(wb, "Top Authors",  ["rank", "author", "publications",
                                     "first_author", "total_citations",
                                     "h_index_estimate", "year_first",
                                     "year_last", "years_active",
                                     "journal_count", "sample_affiliation"],  author_rows)
    _add_sheet(wb, "Top Journals", ["rank", "journal", "abbreviation",
                                     "publications", "percentage",
                                     "total_citations"],                        journal_rows)
    _add_sheet(wb, "Top Countries",["rank", "country", "publications",
                                     "percentage", "total_citations"],          country_rows)
    _add_sheet(wb, "Keywords",     ["rank", "keyword", "frequency"],            kw_rows)
    _add_sheet(wb, "MeSH Terms",   ["rank", "mesh_term", "frequency"],          mesh_rows)
    _add_sheet(wb, "Institutions", ["rank", "institution", "publications"],     inst_rows)
    if lang_rows:
        _add_sheet(wb, "Languages",    ["rank", "code", "language", "n"],          lang_rows)
    if pub_type_rows:
        _add_sheet(wb, "Pub Types",    ["publication_type", "n"],                  pub_type_rows)

    path = pathlib.Path(config.OUTPUT_DIR) / "keratoconus_bibliometrics.xlsx"
    wb.save(path)
    print(f"  wrote: {path.name}")


if __name__ == "__main__":
    data_path = pathlib.Path(config.DATA_DIR) / "analysis.json"
    with open(data_path) as f:
        results = json.load(f)
    print("[report] Generating CSV/Excel reports …")
    generate_reports(results)
    print("[report] Done.")
